#!/usr/bin/env python
"""
测试报销处理功能
"""
import os
import sys

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import django

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "finance_project.settings")
django.setup()

from finance_app.services.excel_processor import ExcelProcessor


def test_reimbursement_processing():
    """测试报销处理功能"""
    processor = ExcelProcessor()

    # 使用报销模板文件进行测试
    test_file = r"d:\kumshing\AutoPython\auto_finance\报销\应用模板 总(1).xlsx"

    try:
        print("开始测试报销处理功能...")
        print(f"测试文件: {test_file}")

        # 处理报销文件
        output_path, record_count = processor.process_excel_file(
            test_file, "reimbursement"
        )

        print(f"处理成功!")
        print(f"输出文件: {output_path}")
        print(f"处理记录数: {record_count}")

        # 检查输出文件是否存在
        if os.path.exists(output_path):
            print("输出文件已成功创建")

            # 简单验证输出文件内容
            import openpyxl

            wb = openpyxl.load_workbook(output_path, data_only=True)
            print(f"输出文件工作表: {wb.sheetnames}")

            if "Page1" in wb.sheetnames:
                sheet = wb["Page1"]
                print(f"数据行数: {sheet.max_row - 1}")  # 减去表头行

            wb.close()
        else:
            print("错误: 输出文件未创建")

    except Exception as e:
        print(f"测试失败: {str(e)}")
        import traceback

        traceback.print_exc()


if __name__ == "__main__":
    test_reimbursement_processing()
