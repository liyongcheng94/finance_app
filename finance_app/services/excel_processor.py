"""
Excel处理服务 - 从原始main.py移植的核心业务逻辑
"""

import os
import datetime
import openpyxl
import xlsxwriter
import json
import logging
from typing import List, Dict, Any, Tuple
from django.conf import settings
from django.contrib.auth.models import User
from ..utils import get_prepared_by_display_name


# 配置常量
BANK_ACCOUNT_MAPPING = {
    "东蜜代深蜜支付：": "1002.16",
    "深蜜代东蜜付：": "1002.21",
    "深蜜支付：": "1002.21",
    "高定支付：": "1002.30.01",
    "高定付：": "1002.30.02",
}

DEFAULT_BANK_ACCOUNT = "1002.16"
DEFAULT_SHEET_NAME = "付款(每日)"

KEY_MAPPINGS = {
    "pay": {
        "日期": "date",
        "支付公司": "company",
        "项目简称（保持一致）": "project",
        "户型": "houseType",
        "供应商": "supplier",
        "费用类型": "feeType",
        "付款方式": "paymentType",
        "总金额": "totalAmount",
        "税": "tax",
        "备注1": "remark1",
        "备注2": "remark2",
        "摘要": "FExplanation",
    },
    "supplier": {
        "简称": "shortName",
        "代码": "code",
        "名称": "name",
        "全名": "fullName",
    },
    "project": {
        "项目简称": "shortName",
        "代码": "code",
        "名称": "name",
        "全名": "fullName",
    },
    "feeType": {"别称": "alias", "科目名称": "name", "科目代码": "code"},
}

EXCEL_HEADERS = [
    "FDate",
    "FYear",
    "FPeriod",
    "FGroupID",
    "FNumber",
    "FAccountNum",
    "FAccountName",
    "FCurrencyNum",
    "FCurrencyName",
    "FAmountFor",
    "FDebit",
    "FCredit",
    "FPreparerID",
    "FCheckerID",
    "FApproveID",
    "FCashierID",
    "FHandler",
    "FSettleTypeID",
    "FSettleNo",
    "FExplanation",
    "FQuantity",
    "FMeasureUnitID",
    "FUnitPrice",
    "FReference",
    "FTransDate",
    "FTransNo",
    "FAttachments",
    "FSerialNum",
    "FObjectName",
    "FParameter",
    "FExchangeRate",
    "FEntryID",
    "FItem",
    "FPosted",
    "FInternalInd",
    "FCashFlow",
]

SCHEMA_HEADERS = [
    "FType",
    "FKey",
    "FFieldName",
    "FCaption",
    "FValueType",
    "FNeedSave",
    "FColIndex",
    "FSrcTableName",
    "FSrcFieldName",
    "FExpFieldName",
    "FImpFieldName",
    "FDefaultVal",
    "FSearch",
    "FItemPageName",
    "FTrueType",
    "FPrecision",
    "FSearchName",
    "FIsShownList",
    "FViewMask",
    "FPage",
]


class ExcelProcessingError(Exception):
    """Excel处理异常"""

    pass


class ExcelProcessor:
    """Excel处理器"""

    def __init__(self):
        self.logger = logging.getLogger("finance_app")

    def get_real_date(self, month_day: Any) -> datetime.datetime:
        """将月日格式转换为完整日期"""
        if month_day is None:
            raise ExcelProcessingError("日期不能为空")

        month_day_str = str(month_day)
        if month_day_str.find(".") == -1:
            raise ExcelProcessingError(f"日期格式错误: {month_day_str}")

        try:
            month = month_day_str.split(".")[0]
            day = month_day_str.split(".")[1]
            now = datetime.datetime.now()
            year = str(now.year)
            result = year + "-" + month + "-" + day
            return datetime.datetime.strptime(result, "%Y-%m-%d")
        except Exception as e:
            raise ExcelProcessingError(f"日期解析错误: {str(e)}")

    def parse_excel(
        self,
        path: str,
        sheet_name: str,
        key_map: Dict[str, str],
        max_col: int = 4,
        start_row: int = 2,
        header_row: int = 1,
    ) -> List[Dict[str, Any]]:
        """解析Excel文件"""
        try:
            result = []
            wb = openpyxl.load_workbook(path, data_only=True)

            if sheet_name not in wb.sheetnames:
                raise ExcelProcessingError(f"工作表 '{sheet_name}' 不存在")

            sheet = wb[sheet_name]

            for row in sheet.iter_rows(
                min_row=start_row, values_only=True, max_col=max_col
            ):
                if row[0] is None:
                    continue
                result.append(row)

            if len(result) == 0:
                self.logger.warning(f"工作表 '{sheet_name}' 中没有数据")
                return []

            # 映射字段名
            mapped_result = []
            header_cells = list(sheet[header_row])

            for i in range(len(result)):
                row = result[i]
                new_row = {}
                for j in range(len(row)):
                    if j >= len(header_cells):
                        break
                    header_value = header_cells[j].value
                    if header_value in key_map:
                        new_row[key_map[header_value]] = row[j]
                mapped_result.append(new_row)

            wb.close()
            return mapped_result

        except Exception as e:
            raise ExcelProcessingError(f"解析Excel文件错误: {str(e)}")

    def supplier_parse_helper(
        self,
        path: str,
        sheet_name: str,
        key_map: Dict[str, str],
        check_list: List[str],
        start_row: int = 2,
        header_row: int = 1,
    ) -> List[Dict[str, Any]]:
        """供应商数据解析辅助函数"""
        try:
            max_col = 4
            result = []
            wb = openpyxl.load_workbook(path, data_only=True)

            if sheet_name not in wb.sheetnames:
                raise ExcelProcessingError(f"工作表 '{sheet_name}' 不存在")

            sheet = wb[sheet_name]

            for row in sheet.iter_rows(
                min_row=start_row, values_only=True, max_col=max_col
            ):
                item = row[0]
                if item is None or item not in check_list:
                    continue
                result.append(row)

            if len(result) == 0:
                self.logger.warning(f"供应商工作表 '{sheet_name}' 中没有匹配的数据")
                return []

            # 映射字段名
            mapped_result = []
            header_cells = list(sheet[header_row])

            for i in range(len(result)):
                row = result[i]
                new_row = {}
                for j in range(len(row)):
                    if j >= len(header_cells):
                        break
                    header_value = header_cells[j].value
                    if header_value in key_map:
                        new_row[key_map[header_value]] = row[j]
                mapped_result.append(new_row)

            wb.close()
            return mapped_result

        except Exception as e:
            raise ExcelProcessingError(f"解析供应商数据错误: {str(e)}")

    def get_real_tax(self, row: Dict[str, Any]) -> float:
        """计算实际税额"""
        payment_type = row.get("paymentType", "")
        remark2 = row.get("remark2", "")
        tax = row.get("tax", 0)
        real_tax = 0

        if payment_type == "partial":
            real_tax = self.get_fee_by_remark2(remark2)
        else:
            real_tax = tax

        if not real_tax:
            real_tax = 0

        return float(real_tax)

    def get_fee_code(self, fee_type_data: List[Dict[str, Any]], fee_type: str) -> str:
        """获取费用代码"""
        for item in fee_type_data:
            if item.get("name") == fee_type:
                return str(item.get("code", ""))
            if fee_type == item.get("name", "") + "费":
                return str(item.get("code", ""))
            if fee_type == item.get("alias", ""):
                return str(item.get("code", ""))
            if fee_type == item.get("alias", "") + "费":
                return str(item.get("code", ""))
        return ""

    def get_fee_by_remark2(self, remark2: str) -> float:
        """从备注2中提取费用"""
        if not remark2:
            raise ExcelProcessingError("请检查有定价的项目 备注2是否填写")

        fee = 0
        try:
            if "（" in remark2 and "）" in remark2:
                remark2 = remark2.replace("（", "").replace("）", "")
                fee = remark2.split("+")[0]
            return float(fee) if fee else 0
        except ValueError:
            raise ExcelProcessingError(f"备注2格式错误: {remark2}")

    def append_data(
        self,
        base_data: List[Dict[str, Any]],
        supplier_data: List[Dict[str, Any]],
        project_data: List[Dict[str, Any]],
        fee_type_data: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        """合并和处理数据"""
        cannot_find = []

        for i in range(len(base_data)):
            row = base_data[i]
            row["supplierStr"] = ""
            row["projectStr"] = ""

            # 查找供应商信息
            for supplier in supplier_data:
                if supplier.get("shortName") == row.get("supplier"):
                    row["supplierStr"] = (
                        "供应商---"
                        + str(supplier.get("code", ""))
                        + "---"
                        + supplier.get("name", "")
                    )
                    break

            # 查找项目信息
            for project in project_data:
                if project.get("shortName") == row.get("project"):
                    row["projectStr"] = (
                        "项目---"
                        + str(project.get("code", ""))
                        + "---"
                        + project.get("name", "")
                    )
                    break

            # 检查是否找到匹配项
            if not row["supplierStr"]:
                error_string = f"找不到供应商: {row.get('supplier', '')}"
                cannot_find.append(error_string)

            if not row["projectStr"]:
                cannot_find.append(f"找不到项目: {row.get('project', '')}")

            if len(cannot_find) > 0:
                continue

            row["depStr"] = "部门---02---采购部"
            row["depProjectStr"] = row["depStr"] + "," + row["projectStr"]
            row["supplierProjectStr"] = row["supplierStr"] + "," + row["projectStr"]

            # 设置银行账户
            pay_company = row.get("company", "")
            row["FAccountNum_Bank"] = BANK_ACCOUNT_MAPPING.get(
                pay_company, DEFAULT_BANK_ACCOUNT
            )

            row["feeCode"] = self.get_fee_code(fee_type_data, row.get("feeType", ""))
            real_date = self.get_real_date(row.get("date"))
            row["FDate"] = real_date.strftime("%Y-%m-%d")
            row["FYear"] = real_date.strftime("%Y")
            row["FPeriod"] = real_date.strftime("%m")

            payment_type = row.get("paymentType", "")
            if "定金" in payment_type:
                row["paymentType"] = "partial"
            if "全款" in payment_type:
                row["paymentType"] = "full"

            real_tax = self.get_real_tax(row)
            row["realTax"] = real_tax
            row["remain"] = float(row.get("totalAmount", 0)) - real_tax

        if len(cannot_find) > 0:
            raise ExcelProcessingError(
                f"请检查项目或供应商是否正确: {', '.join(cannot_find)}"
            )

        return base_data

    def gen_excel_row(
        self, row: Dict[str, Any], index: int, user: User = None
    ) -> List[Dict[str, Any]]:
        """生成Excel行数据"""
        # 通用数据
        common_data = {
            "FDate": row["FDate"],
            "FYear": row["FYear"],
            "FPeriod": row["FPeriod"],
            "FGroupID": "记",
            "FNumber": index,
            "FCurrencyNum": "RMB",
            "FCurrencyName": "人民币",
            "FPreparerID": get_prepared_by_display_name(user),
            "FCheckerID": "NONE",
            "FApproveID": "NONE",
            "FCashierID": "NONE",
            "FHandler": "",
            "FSettleTypeID": "*",
            "FSettleNo": "",
            "FExplanation": row.get("FExplanation", ""),
            "FQuantity": 0,
            "FMeasureUnitID": "*",
            "FUnitPrice": 0,
            "FReference": "",
            "FTransDate": row["FDate"],
            "FTransNo": "",
            "FAttachments": 0,
            "FSerialNum": index,
            "FObjectName": "",
            "FParameter": "",
            "FExchangeRate": 1,
            "FPosted": 0,
            "FInternalInd": "",
            "FCashFlow": "",
        }

        payment_type = row.get("paymentType", "")
        real_tax = row.get("realTax", 0)
        total_amount = float(row.get("totalAmount", 0))
        remain = row.get("remain", 0)

        # 根据支付类型生成不同的行数据
        if payment_type == "partial":
            unique_data = [
                {
                    "FAccountNum": row.get("feeCode", ""),
                    "FAccountName": row.get("feeType", ""),
                    "FAmountFor": real_tax,
                    "FDebit": real_tax,
                    "FCredit": 0,
                    "FEntryID": 0,
                    "FItem": row.get("depProjectStr", ""),
                },
                {
                    "FAccountNum": "1402",
                    "FAccountName": "在途物资",
                    "FAmountFor": remain,
                    "FDebit": remain,
                    "FCredit": 0,
                    "FEntryID": 1,
                    "FItem": "",
                },
                {
                    "FAccountNum": "2202.01",
                    "FAccountName": "应付供应商",
                    "FAmountFor": total_amount,
                    "FDebit": 0,
                    "FCredit": total_amount,
                    "FEntryID": 2,
                    "FItem": row.get("supplierProjectStr", ""),
                },
                {
                    "FAccountNum": "2202.01",
                    "FAccountName": "应付供应商",
                    "FAmountFor": real_tax,
                    "FDebit": real_tax,
                    "FCredit": 0,
                    "FEntryID": 3,
                    "FItem": row.get("supplierProjectStr", ""),
                },
                {
                    "FAccountNum": row.get("FAccountNum_Bank", ""),
                    "FAccountName": "银行",
                    "FAmountFor": real_tax,
                    "FDebit": 0,
                    "FCredit": real_tax,
                    "FEntryID": 4,
                    "FItem": "",
                },
            ]
        elif payment_type == "full":
            if real_tax == 0:
                # 全款不含税
                unique_data = [
                    {
                        "FAccountNum": row.get("feeCode", ""),
                        "FAccountName": row.get("feeType", ""),
                        "FAmountFor": remain,
                        "FDebit": remain,
                        "FCredit": 0,
                        "FEntryID": 0,
                        "FItem": row.get("depProjectStr", ""),
                    },
                    {
                        "FAccountNum": "2202.01",
                        "FAccountName": "应付供应商",
                        "FAmountFor": total_amount,
                        "FDebit": 0,
                        "FCredit": total_amount,
                        "FEntryID": 1,
                        "FItem": row.get("supplierProjectStr", ""),
                    },
                    {
                        "FAccountNum": "2202.01",
                        "FAccountName": "应付供应商",
                        "FAmountFor": total_amount,
                        "FDebit": total_amount,
                        "FCredit": 0,
                        "FEntryID": 2,
                        "FItem": row.get("supplierProjectStr", ""),
                    },
                    {
                        "FAccountNum": row.get("FAccountNum_Bank", ""),
                        "FAccountName": "银行",
                        "FAmountFor": total_amount,
                        "FDebit": 0,
                        "FCredit": total_amount,
                        "FEntryID": 3,
                        "FItem": "",
                    },
                ]
            else:
                # 全款含税
                unique_data = [
                    {
                        "FAccountNum": row.get("feeCode", ""),
                        "FAccountName": row.get("feeType", ""),
                        "FAmountFor": remain,
                        "FDebit": remain,
                        "FCredit": 0,
                        "FEntryID": 0,
                        "FItem": row.get("depProjectStr", ""),
                    },
                    {
                        "FAccountNum": "2221.01.01",
                        "FAccountName": "进项税额",
                        "FAmountFor": real_tax,
                        "FDebit": real_tax,
                        "FCredit": 0,
                        "FEntryID": 1,
                        "FItem": "",
                    },
                    {
                        "FAccountNum": "2202.01",
                        "FAccountName": "应付供应商",
                        "FAmountFor": total_amount,
                        "FDebit": 0,
                        "FCredit": total_amount,
                        "FEntryID": 2,
                        "FItem": row.get("supplierProjectStr", ""),
                    },
                    {
                        "FAccountNum": "2202.01",
                        "FAccountName": "应付供应商",
                        "FAmountFor": total_amount,
                        "FDebit": total_amount,
                        "FCredit": 0,
                        "FEntryID": 3,
                        "FItem": row.get("supplierProjectStr", ""),
                    },
                    {
                        "FAccountNum": row.get("FAccountNum_Bank", ""),
                        "FAccountName": "银行",
                        "FAmountFor": total_amount,
                        "FDebit": 0,
                        "FCredit": total_amount,
                        "FEntryID": 4,
                        "FItem": "",
                    },
                ]
        else:
            unique_data = []

        # 合并数据
        result = []
        for item in unique_data:
            result.append({**item, **common_data})

        return result

    def gen_excel_data(
        self, data: List[Dict[str, Any]], user: User = None
    ) -> List[List[Dict[str, Any]]]:
        """生成所有Excel数据"""
        excel_data = []
        index = 0

        for row in data:
            index = index + 1
            excel_data.append(self.gen_excel_row(row, index, user))

        return excel_data

    def write_excel(self, excel_file: str, data: List[List[Dict[str, Any]]]) -> None:
        """写入Excel文件"""
        try:
            workbook = xlsxwriter.Workbook(excel_file)

            # 写入schema工作表
            worksheet = workbook.add_worksheet("t_Schema")

            # 尝试读取schema文件
            schema_file_path = os.path.join(
                os.path.dirname(__file__), "..", "..", "t_Schema.json"
            )
            try:
                with open(schema_file_path, "r", encoding="utf-8") as f:
                    schema = json.load(f)
            except FileNotFoundError:
                self.logger.warning("t_Schema.json 文件未找到，使用空schema")
                schema = []

            # 写入schema表头
            for i, header in enumerate(SCHEMA_HEADERS):
                worksheet.write(0, i, header)

            # 写入schema数据
            for i, schema_item in enumerate(schema):
                for j, header in enumerate(SCHEMA_HEADERS):
                    if header in schema_item:
                        worksheet.write(i + 1, j, schema_item[header])

            # 写入数据工作表
            worksheet = workbook.add_worksheet("Page1")

            # 写入数据表头
            for i, header in enumerate(EXCEL_HEADERS):
                worksheet.write(0, i, header)

            # 写入数据
            sum_row = 0
            for i in range(len(data)):
                for j in range(len(data[i])):
                    sum_row = sum_row + 1
                    for k, header in enumerate(EXCEL_HEADERS):
                        if header in data[i][j]:
                            worksheet.write(sum_row, k, data[i][j][header])

            workbook.close()

        except Exception as e:
            raise ExcelProcessingError(f"写入Excel文件错误: {str(e)}")

    def process_excel_file(
        self,
        file_path: str,
        process_type: str = "payment",
        user: User = None,
    ) -> Tuple[str, int]:
        """
        处理Excel文件的主要方法

        Args:
            file_path: 输入Excel文件路径
            process_type: 处理类型 ("payment" 或 "reimbursement")
            user: Django用户对象，用于获取制单人信息

        Returns:
            tuple: (输出文件路径, 处理的记录数)
        """
        self.logger.info(f"开始处理Excel文件: {file_path}, 处理类型: {process_type}")

        if process_type == "reimbursement":
            return self.process_reimbursement_file(file_path, user)
        else:
            return self.process_payment_file(file_path, user)

    def process_payment_file(
        self, file_path: str, user: User = None
    ) -> Tuple[str, int]:
        """处理排单文件"""
        self.logger.info(f"开始处理排单Excel文件: {file_path}")

        try:
            # 解析基础数据
            base_data = self.parse_excel(
                file_path,
                DEFAULT_SHEET_NAME,
                key_map=KEY_MAPPINGS["pay"],
                max_col=12,
                start_row=3,
                header_row=2,
            )

            if not base_data:
                raise ExcelProcessingError("基础数据为空")

            # 提取供应商列表
            supplier_check_list = [
                item.get("supplier", "") for item in base_data if item.get("supplier")
            ]

            # 解析辅助数据
            fee_type_data = self.parse_excel(
                file_path,
                "核算项目_费用代码",
                max_col=3,
                key_map=KEY_MAPPINGS["feeType"],
            )
            project_data = self.parse_excel(
                file_path, "核算项目_项目", key_map=KEY_MAPPINGS["project"]
            )
            supplier_data = self.supplier_parse_helper(
                file_path,
                "核算项目_供应商",
                key_map=KEY_MAPPINGS["supplier"],
                check_list=supplier_check_list,
            )

            # 处理数据
            processed_data = self.append_data(
                base_data, supplier_data, project_data, fee_type_data
            )
            excel_data = self.gen_excel_data(processed_data, user)

            # 生成输出文件路径
            now = datetime.datetime.now()
            timestamp = now.strftime("%Y%m%d_%H%M%S")
            output_filename = f"排单_{timestamp}.xlsx"
            output_path = os.path.join(settings.MEDIA_ROOT, "outputs", output_filename)

            # 确保输出目录存在
            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            # 写入Excel文件
            self.write_excel(output_path, excel_data)

            self.logger.info(f"排单Excel文件处理完成: {output_path}")
            return output_path, len(processed_data)

        except Exception as e:
            self.logger.error(f"处理排单Excel文件时发生错误: {str(e)}")
            raise ExcelProcessingError(f"处理排单Excel文件失败: {str(e)}")

    def process_reimbursement_file(
        self, file_path: str, user: User = None
    ) -> Tuple[str, int]:
        """处理报销文件"""
        self.logger.info(f"开始处理报销Excel文件: {file_path}")

        try:
            # 解析费用代码映射表
            fee_type_mapping = self.parse_reimbursement_fee_mapping(file_path)

            # 解析顶部信息（报销人、日期、银行等）
            top_infos = self.parse_reimbursement_top_info(file_path)

            # 解析基础报销数据
            base_data = self.parse_reimbursement_base_data(file_path, len(top_infos))

            # 处理报销数据，按报销人分组生成会计分录
            result_data = self.process_reimbursement_data(
                top_infos, base_data, fee_type_mapping, user
            )

            # 生成输出文件路径
            now = datetime.datetime.now()
            timestamp = now.strftime("%Y%m%d_%H%M%S")
            output_filename = f"报销_{timestamp}.xlsx"
            output_path = os.path.join(settings.MEDIA_ROOT, "outputs", output_filename)

            # 确保输出目录存在
            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            # 写入Excel文件
            self.write_reimbursement_excel(output_path, result_data)

            self.logger.info(f"报销Excel文件处理完成: {output_path}")
            return output_path, len(result_data)

        except Exception as e:
            self.logger.error(f"处理报销Excel文件时发生错误: {str(e)}")
            raise ExcelProcessingError(f"处理报销Excel文件失败: {str(e)}")

    def parse_reimbursement_fee_mapping(self, file_path: str) -> Dict[str, str]:
        """解析报销文件中的费用代码映射表"""
        try:
            mapping = {}
            wb = openpyxl.load_workbook(file_path, data_only=True)

            # 读取主费用代码表
            sheet_name = "核算项目_费用代码"
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                # 解析费用代码映射关系：费用代码 -> 科目名称
                for row in sheet.iter_rows(
                    min_row=2, values_only=True, max_col=3
                ):  # 跳过表头
                    if row[2] and row[1]:  # 费用代码和科目名称都不为空
                        fee_code = str(row[2]).strip()
                        fee_name = str(row[1]).strip()
                        mapping[fee_code] = fee_name

            # 读取无项目费用代码表
            sheet_name_no_project = "核算项目_费用代码无项目"
            if sheet_name_no_project in wb.sheetnames:
                sheet = wb[sheet_name_no_project]
                # 解析费用代码映射关系：费用代码 -> 科目名称
                for row in sheet.iter_rows(
                    min_row=2, values_only=True, max_col=3
                ):  # 跳过表头
                    if row[2] and row[1]:  # 费用代码和科目名称都不为空
                        fee_code = str(row[2]).strip()
                        fee_name = str(row[1]).strip()
                        # 如果已经存在，不覆盖（优先使用主表的数据）
                        if fee_code not in mapping:
                            mapping[fee_code] = fee_name

            wb.close()
            self.logger.info(f"成功加载费用代码映射表，共 {len(mapping)} 项")
            return mapping

        except Exception as e:
            self.logger.warning(
                f"解析费用代码映射表错误: {str(e)}，将使用费用别称作为费用类型"
            )
            return {}

    def parse_reimbursement_top_info(self, file_path: str) -> List[Dict[str, Any]]:
        """解析报销文件顶部信息"""
        try:
            result = []
            wb = openpyxl.load_workbook(file_path, data_only=True)

            # 报销文件的工作表名
            sheet_name = "报销  (每日)"
            if sheet_name not in wb.sheetnames:
                raise ExcelProcessingError(f"工作表 '{sheet_name}' 不存在")

            sheet = wb[sheet_name]

            # 解析顶部信息行（以"日期"开头的行）
            for row in sheet.iter_rows(min_row=1, values_only=True, max_col=16):
                if row[0] == "日期":
                    data = {
                        "date": row[1],
                        "person": row[3],
                        "bank": row[5],
                        "bankCode": row[7],
                        "summary": row[9],
                    }
                    result.append(data)
                else:
                    if result:  # 如果已经找到了数据，遇到非"日期"行就停止
                        break

            wb.close()
            return result

        except Exception as e:
            raise ExcelProcessingError(f"解析报销顶部信息错误: {str(e)}")

    def parse_reimbursement_base_data(
        self, file_path: str, top_count: int
    ) -> List[Dict[str, Any]]:
        """解析报销基础数据"""
        try:
            result = []
            wb = openpyxl.load_workbook(file_path, data_only=True)

            sheet_name = "报销  (每日)"
            sheet = wb[sheet_name]

            # 从第4行开始解析基础数据（第3行是表头）
            start_row = 4

            for row in sheet.iter_rows(min_row=start_row, values_only=True, max_col=16):
                # 解析报销明细：报销人、部门代码、项目简称、备注、费用别称、备注、费用别称、费用类型、费用代码、金额、摘要、部门+项目
                if row[0] is None:  # 如果报销人为空，跳过
                    continue

                dic = {
                    "person": row[0],  # 报销人
                    "department": row[1],  # 部门代码
                    "project": row[2],  # 项目简称
                    "remark": row[3],  # 备注
                    "feeAlias": row[4],  # 费用别称
                    "feeType": row[4],  # 费用类型（使用费用别称）
                    "feeCode": row[6],  # 费用代码
                    "amount": row[7],  # 金额
                    "summary": row[8],  # 摘要
                    "departmentProject": row[9] if len(row) > 9 else "",  # 部门+项目
                }

                # 如果关键字段都为空，跳过这行
                if (dic["person"] is None or dic["person"] == "") and (
                    dic["amount"] is None or dic["amount"] == 0
                ):
                    continue

                # 检查部门+项目是否为空或出错
                dept_project = dic["departmentProject"]
                if (
                    dept_project is None
                    or dept_project == ""
                    or str(dept_project).startswith("#")
                ):
                    if dic["project"] is not None and dic["project"] != "":
                        # 如果项目不为空但部门+项目为空，给出警告但继续处理
                        self.logger.warning(
                            f"报销人 {dic['person']} 的项目 {dic['project']} 部门+项目字段为空或有错误: {dept_project}"
                        )
                        dic["departmentProject"] = ""  # 设为空字符串以便后续处理

                result.append(dic)

            wb.close()
            return result

        except Exception as e:
            raise ExcelProcessingError(f"解析报销基础数据错误: {str(e)}")

    def process_reimbursement_data(
        self,
        top_infos: List[Dict[str, Any]],
        base_data: List[Dict[str, Any]],
        fee_type_mapping: Dict[str, str],
        user: User = None,
    ) -> List[Dict[str, Any]]:
        """处理报销数据，生成会计分录"""
        try:
            result = []

            if not top_infos:
                raise ExcelProcessingError("未找到报销人信息")

            date = top_infos[0]["date"]
            person_index = 1

            # 按报销人分组处理
            for top_info in top_infos:
                index = 0
                total_amount = 0
                person = top_info["person"]

                # 处理该报销人的所有明细
                for item in base_data:
                    if (
                        item["person"] is None
                        or item["person"] == ""
                        or person != item["person"]
                    ):
                        continue

                    # 确保金额格式正确
                    amount = float(item["amount"]) if item["amount"] else 0
                    item["amount"] = round(amount, 2)
                    total_amount += item["amount"]
                    total_amount = round(total_amount, 2)

                    # 生成费用明细的会计分录（借方）
                    single_data = self.gen_reimbursement_debit_row(
                        user, item, index, date, person_index, fee_type_mapping
                    )
                    result.extend(single_data)
                    index += 1

                # 生成银行支付的会计分录（贷方）
                if total_amount > 0:
                    last_row = self.gen_reimbursement_credit_row(
                        user,
                        top_info,
                        base_data,
                        index,
                        date,
                        person_index,
                        total_amount,
                    )
                    result.extend(last_row)

                person_index += 1

            return result

        except Exception as e:
            raise ExcelProcessingError(f"处理报销数据错误: {str(e)}")

    def gen_reimbursement_debit_row(
        self,
        user: Any,
        item: Dict[str, Any],
        index: int,
        date: Any,
        person_index: int,
        fee_type_mapping: Dict[str, str],
    ) -> List[Dict[str, Any]]:
        """生成报销费用的借方分录"""
        real_date = self.get_real_date(date)

        # 基础数据
        base_info = {
            "FDate": real_date.strftime("%Y-%m-%d"),
            "FYear": real_date.strftime("%Y"),
            "FPeriod": real_date.strftime("%m"),
            "FGroupID": "记",
            "FNumber": person_index,
            "FCurrencyNum": "RMB",
            "FCurrencyName": "人民币",
            "FPreparerID": get_prepared_by_display_name(user),
            "FCheckerID": "NONE",
            "FApproveID": "NONE",
            "FCashierID": "NONE",
            "FHandler": "",
            "FSettleTypeID": "*",
            "FSettleNo": "",
            "FQuantity": 0,
            "FMeasureUnitID": "*",
            "FUnitPrice": 0,
            "FReference": "",
            "FTransDate": real_date.strftime("%Y-%m-%d"),
            "FTransNo": "",
            "FAttachments": 0,
            "FSerialNum": 1,
            "FObjectName": "",
            "FParameter": "",
            "FExchangeRate": 1,
            "FPosted": 0,
            "FInternalInd": "",
            "FCashFlow": "",
        }

        # 获取正确的费用类型名称
        fee_code = str(item["feeCode"]) if item["feeCode"] else ""
        fee_name = fee_type_mapping.get(
            fee_code, item["feeType"] if item["feeType"] else ""
        )

        # 费用明细行（借方）
        debit_row = {
            **base_info,
            "FAccountNum": fee_code,
            "FAccountName": fee_name,
            "FAmountFor": item["amount"],
            "FDebit": item["amount"],
            "FCredit": 0,
            "FExplanation": str(date) + (item["summary"] if item["summary"] else ""),
            "FEntryID": index,
            "FItem": item["departmentProject"] if item["departmentProject"] else "",
        }

        return [debit_row]

    def gen_reimbursement_credit_row(
        self,
        user: Any,
        top_info: Dict[str, Any],
        base_data: List[Dict[str, Any]],
        index: int,
        date: Any,
        person_index: int,
        total_amount: float,
    ) -> List[Dict[str, Any]]:
        """生成银行支付的贷方分录"""
        real_date = self.get_real_date(date)

        # 基础数据
        base_info = {
            "FDate": real_date.strftime("%Y-%m-%d"),
            "FYear": real_date.strftime("%Y"),
            "FPeriod": real_date.strftime("%m"),
            "FGroupID": "记",
            "FNumber": person_index,
            "FCurrencyNum": "RMB",
            "FCurrencyName": "人民币",
            "FPreparerID": get_prepared_by_display_name(user),
            "FCheckerID": "NONE",
            "FApproveID": "NONE",
            "FCashierID": "NONE",
            "FHandler": "",
            "FSettleTypeID": "*",
            "FSettleNo": "",
            "FQuantity": 0,
            "FMeasureUnitID": "*",
            "FUnitPrice": 0,
            "FReference": "",
            "FTransDate": real_date.strftime("%Y-%m-%d"),
            "FTransNo": "",
            "FAttachments": 0,
            "FSerialNum": 1,
            "FObjectName": "",
            "FParameter": "",
            "FExchangeRate": 1,
            "FPosted": 0,
            "FInternalInd": "",
            "FCashFlow": "",
        }

        # 银行支付行（贷方）
        credit_row = {
            **base_info,
            "FAccountNum": top_info["bankCode"],
            "FAccountName": top_info["bank"],
            "FAmountFor": total_amount,
            "FDebit": 0,
            "FCredit": total_amount,
            "FExplanation": str(date) + top_info["summary"],
            "FEntryID": index,
            "FItem": "",
        }

        return [credit_row]

    def write_reimbursement_excel(
        self, excel_file: str, data: List[Dict[str, Any]]
    ) -> None:
        """写入报销Excel文件"""
        try:
            workbook = xlsxwriter.Workbook(excel_file)

            # 写入schema工作表
            worksheet = workbook.add_worksheet("t_Schema")

            # 尝试读取schema文件
            schema_file_path = os.path.join(
                os.path.dirname(__file__), "..", "..", "t_Schema.json"
            )
            try:
                with open(schema_file_path, "r", encoding="utf-8") as f:
                    schema = json.load(f)
            except FileNotFoundError:
                self.logger.warning("t_Schema.json 文件未找到，使用空schema")
                schema = []

            # 写入schema表头
            for i, header in enumerate(SCHEMA_HEADERS):
                worksheet.write(0, i, header)

            # 写入schema数据
            for i, schema_item in enumerate(schema):
                for j, header in enumerate(SCHEMA_HEADERS):
                    if header in schema_item:
                        worksheet.write(i + 1, j, schema_item[header])

            # 写入数据工作表
            worksheet = workbook.add_worksheet("Page1")

            # 写入数据表头
            for i, header in enumerate(EXCEL_HEADERS):
                worksheet.write(0, i, header)

            # 写入数据
            for i, row_data in enumerate(data):
                for j, header in enumerate(EXCEL_HEADERS):
                    if header == "FAmountFor":
                        # FAmountFor列使用求和公式
                        worksheet.write(i + 1, j, f"=SUM(K{i + 2}+L{i + 2})")
                    elif header in row_data:
                        worksheet.write(i + 1, j, row_data[header])

            workbook.close()

        except Exception as e:
            raise ExcelProcessingError(f"写入报销Excel文件错误: {str(e)}")
