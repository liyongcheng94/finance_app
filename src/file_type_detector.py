"""
工作表检测器 - 分析Excel文件中的工作表信息
"""

import openpyxl
import sys
import os
from typing import Dict, List, Optional

# 添加路径以便导入配置
sys.path.append(os.path.dirname(__file__))


class SheetAnalyzer:
class SheetAnalyzer:
    """工作表分析器 - 分析Excel文件中的工作表信息"""

    def __init__(self):
        self.workbook = None

    def analyze_file(self, file_path: str) -> Dict:
        """
        分析Excel文件，返回工作表信息
        参数: file_path - Excel文件路径
        返回: 包含工作表信息的字典
        """
        try:
            self.workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            sheet_info = {
                'file_path': file_path,
                'sheet_names': self.workbook.sheetnames,
                'sheet_details': {},
                'suggested_types': {}
            }
            
            # 分析每个工作表
            for sheet_name in self.workbook.sheetnames:
                details = self._analyze_sheet(sheet_name)
                sheet_info['sheet_details'][sheet_name] = details
                sheet_info['suggested_types'][sheet_name] = self._suggest_sheet_type(details)
            
            return sheet_info
            
        except Exception as e:
            return {
                'error': f"文件分析失败: {e}",
                'file_path': file_path,
                'sheet_names': [],
                'sheet_details': {},
                'suggested_types': {}
            }
        finally:
            if self.workbook:
                self.workbook.close()

    def _analyze_sheet(self, sheet_name: str) -> Dict:
        """分析单个工作表"""
        try:
            sheet = self.workbook[sheet_name]
            
            # 获取工作表基本信息
            details = {
                'name': sheet_name,
                'max_row': sheet.max_row,
                'max_col': sheet.max_column,
                'headers': [],
                'sample_data': [],
                'keywords_found': []
            }
            
            # 分析表头（前3行）
            for row_idx in range(1, min(4, sheet.max_row + 1)):
                row_data = []
                for cell in sheet[row_idx]:
                    value = str(cell.value) if cell.value else ""
                    row_data.append(value)
                
                if row_idx == 1:
                    details['headers'] = row_data[:10]  # 只取前10列作为表头
                else:
                    details['sample_data'].append(row_data[:10])
            
            # 查找关键词
            all_text = " ".join(details['headers'] + [str(item) for sublist in details['sample_data'] for item in sublist])
            
            payment_keywords = ["支付公司", "供应商", "付款方式", "户型", "排单"]
            reimbursement_keywords = ["报销人", "部门代码", "费用类型", "费用代码", "报销", "银行"]
            
            for keyword in payment_keywords:
                if keyword in all_text:
                    details['keywords_found'].append(f"排单关键词: {keyword}")
            
            for keyword in reimbursement_keywords:
                if keyword in all_text:
                    details['keywords_found'].append(f"报销关键词: {keyword}")
            
            return details
            
        except Exception as e:
            return {
                'name': sheet_name,
                'error': str(e),
                'headers': [],
                'sample_data': [],
                'keywords_found': []
            }

    def _suggest_sheet_type(self, sheet_details: Dict) -> str:
        """根据工作表内容建议类型"""
        keywords = " ".join(sheet_details.get('keywords_found', []))
        
        if "报销关键词" in keywords:
            return "reimbursement"
        elif "排单关键词" in keywords:
            return "payment"
        else:
            return "unknown"

    def get_sheet_names(self, file_path: str) -> List[str]:
        """快速获取工作表名称列表"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
            return sheet_names
        except Exception as e:
            print(f"获取工作表名称失败: {e}")
            return []
