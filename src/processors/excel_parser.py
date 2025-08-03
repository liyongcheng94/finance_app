"""
Excel文件解析和处理功能
"""

import openpyxl
import logging
import os
import sys

sys.path.append(os.path.dirname(os.path.dirname(__file__)))

from utils.helpers import exit_with_message


class ExcelParser:
    """Excel文件解析器"""

    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None

    def load_workbook(self):
        """加载Excel工作簿"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            return self.workbook
        except Exception as e:
            exit_with_message(f"加载Excel文件失败: {str(e)}")

    def parse_sheet(self, sheet_name, key_map, max_col=4, start_row=2, header_row=1):
        """解析Excel工作表"""
        try:
            result = []

            if not self.workbook:
                self.load_workbook()

            # 获取工作表
            sheet = self.workbook[sheet_name]

            # 读取数据行
            for row in sheet.iter_rows(
                min_row=start_row, values_only=True, max_col=max_col
            ):
                if row[0] is None:
                    continue
                result.append(row)

            if len(result) == 0:
                logging.error("未找到数据")
                return []

            # 映射字段名
            mapped_result = []
            for i in range(len(result)):
                row = result[i]
                new_row = {}
                for j in range(len(row)):
                    if j >= len(sheet[header_row]):
                        break
                    header_value = sheet[header_row][j].value
                    if header_value in key_map:
                        new_row[key_map[header_value]] = row[j]
                mapped_result.append(new_row)

            return mapped_result

        except Exception as e:
            exit_with_message(f"解析Excel失败: {str(e)}")

    def parse_supplier_helper(
        self, sheet_name, key_map, check_list, start_row=2, header_row=1
    ):
        """供应商数据解析辅助函数"""
        try:
            max_col = 4
            result = []

            if not self.workbook:
                self.load_workbook()

            sheet = self.workbook[sheet_name]

            for row in sheet.iter_rows(
                min_row=start_row, values_only=True, max_col=max_col
            ):
                item = row[0]
                if item is None or item not in check_list:
                    continue
                result.append(row)

            if len(result) == 0:
                logging.error("未找到供应商数据")
                return []

            # 映射字段名
            mapped_result = []
            for i in range(len(result)):
                row = result[i]
                new_row = {}
                for j in range(len(row)):
                    if j >= len(sheet[header_row]):
                        break
                    header_value = sheet[header_row][j].value
                    if header_value in key_map:
                        new_row[key_map[header_value]] = row[j]
                mapped_result.append(new_row)

            return mapped_result

        except Exception as e:
            exit_with_message(f"解析供应商数据失败: {str(e)}")

    def close(self):
        """关闭工作簿"""
        if self.workbook:
            self.workbook.close()
