"""
报销系统处理器 - 统一接口
"""

import openpyxl
import logging
from typing import List, Optional, Dict, Any
import datetime
import sys
import os

# 添加路径以便导入其他模块
sys.path.append(os.path.dirname(os.path.dirname(__file__)))

from models.reimbursement_models import (
    ReimbursementTopInfo,
    ReimbursementData,
    ReimbursementExcelRowData,
)
from unified_config import REIMBURSEMENT_CONFIG
from config import DEFAULT_USERS


class ReimbursementProcessor:
    """报销数据处理器 - 统一接口"""

    def __init__(self):
        self.parser = ReimbursementExcelParser()
        self.data_processor = ReimbursementDataProcessor()

    def process_file(self, file_path: str) -> List[Dict[str, Any]]:
        """处理报销文件的统一接口"""
        try:
            # 解析Excel数据
            top_infos = self.parser.get_top_infos(file_path)
            base_data = self.parser.get_base_data(file_path)

            if not top_infos:
                raise Exception("未找到顶部信息数据")

            if not base_data:
                raise Exception("未找到基础数据")

            # 处理数据
            result = self.data_processor.process_reimbursement_data(
                top_infos, base_data
            )
            return result

        except Exception as e:
            raise Exception(f"处理报销文件失败: {e}")


def get_real_date(month_day):
    """将月.日格式转换为完整日期"""
    if month_day is None:
        return None

    # 解析monthDay为字符串
    month_day_str = str(month_day)
    if month_day_str.find(".") == -1:
        return None

    month = month_day_str.split(".")[0]
    day = month_day_str.split(".")[1]

    # 获取当前年份
    now = datetime.datetime.now()
    year = str(now.year)
    result = year + "-" + month + "-" + day
    return datetime.datetime.strptime(result, "%Y-%m-%d")


def round_amount(amount, decimals=2):
    """金额四舍五入到指定小数位"""
    try:
        return round(float(amount), decimals)
    except (ValueError, TypeError):
        return 0.0


class ReimbursementExcelParser:
    """报销Excel文件解析器"""

    def __init__(self):
        self.sheet_name = REIMBURSEMENT_CONFIG["DEFAULT_SHEET_NAME"]
        self.workbook = None
        self.sheet = None
        self.top_index = 0

    def load_workbook(self, file_path: str):
        """加载Excel工作簿"""
        try:
            self.workbook = openpyxl.load_workbook(file_path, data_only=True)
            self.sheet = self.workbook[self.sheet_name]
            return True
        except Exception as e:
            raise Exception(f"加载Excel文件失败: {str(e)}")

    def get_top_infos(self, file_path: str) -> List[ReimbursementTopInfo]:
        """获取顶部信息"""
        if not self.load_workbook(file_path):
            return []

        result = []
        TOP_INFO_COLUMNS = REIMBURSEMENT_CONFIG["TOP_INFO_COLUMNS"]

        for row in self.sheet.iter_rows(min_row=0, values_only=True, max_col=16):
            if row[0] == "日期":
                top_info = ReimbursementTopInfo(
                    date=row[TOP_INFO_COLUMNS["date"]],
                    person=row[TOP_INFO_COLUMNS["person"]],
                    bank=row[TOP_INFO_COLUMNS["bank"]],
                    bank_code=row[TOP_INFO_COLUMNS["bankCode"]],
                    summary=row[TOP_INFO_COLUMNS["summary"]],
                )
                self.top_index += 1
                result.append(top_info)
            else:
                break

        return result

    def get_base_data(self, file_path: str) -> List[ReimbursementData]:
        """获取基础数据"""
        if not self.load_workbook(file_path):
            return []

        result = []
        REIMBURSEMENT_COLUMNS = REIMBURSEMENT_CONFIG["REIMBURSEMENT_COLUMNS"]

        # 从第3+topIndex行开始读取数据
        for row in self.sheet.iter_rows(
            min_row=3 + self.top_index, values_only=True, max_col=16
        ):
            # 创建报销数据对象
            reimbursement_data = ReimbursementData(
                person=row[REIMBURSEMENT_COLUMNS["person"]],
                department=row[REIMBURSEMENT_COLUMNS["department"]],
                project=row[REIMBURSEMENT_COLUMNS["project"]],
                remark=row[REIMBURSEMENT_COLUMNS["remark"]],
                fee_type=row[REIMBURSEMENT_COLUMNS["feeType"]],
                fee_code=row[REIMBURSEMENT_COLUMNS["feeCode"]],
                amount=row[REIMBURSEMENT_COLUMNS["amount"]],
                summary=row[REIMBURSEMENT_COLUMNS["summary"]],
                department_project=row[REIMBURSEMENT_COLUMNS["departmentProject"]],
            )

            # 跳过空行
            if reimbursement_data.is_empty():
                continue

            # 处理项目为空的情况
            if reimbursement_data.project is None or reimbursement_data.project == "":
                reimbursement_data.department_project = ""

            # 验证部门+项目字段
            if (
                reimbursement_data.department_project is None
                or reimbursement_data.department_project == ""
            ):
                continue

            result.append(reimbursement_data)

        return result


class ReimbursementDataProcessor:
    """报销数据处理器"""

    def __init__(self):
        pass

    def generate_excel_single_data(
        self,
        reimbursement_data: ReimbursementData,
        index: int,
        date: str,
        person_index: int,
    ) -> List[Dict[str, Any]]:
        """生成单条Excel数据"""
        result = []
        real_date = get_real_date(date)

        if not real_date:
            return result

        # 创建Excel行数据
        excel_row = ReimbursementExcelRowData(
            f_date=real_date.strftime("%Y-%m-%d"),
            f_year=real_date.strftime("%Y"),
            f_period=real_date.strftime("%m"),
            f_group_id="记",
            f_number=person_index,
            f_account_num=reimbursement_data.fee_code,
            f_account_name=reimbursement_data.fee_type,
            f_currency_num="RMB",
            f_currency_name="人民币",
            f_amount_for=reimbursement_data.amount,
            f_debit=reimbursement_data.amount,
            f_credit=0,
            f_preparer_id=DEFAULT_USERS["PREPARER"],
            f_checker_id="NONE",
            f_approve_id="NONE",
            f_cashier_id="NONE",
            f_handler="",
            f_settle_type_id="*",
            f_settle_no="",
            f_explanation=str(date) + str(reimbursement_data.summary or ""),
            f_quantity=0,
            f_measure_unit_id="*",
            f_unit_price=0,
            f_reference="",
            f_trans_date=real_date.strftime("%Y-%m-%d"),
            f_trans_no="",
            f_attachments=0,
            f_serial_num=1,
            f_object_name="",
            f_parameter="",
            f_exchange_rate=1,
            f_entry_id=index,
            f_item=reimbursement_data.department_project,
            f_posted=0,
            f_internal_ind="",
            f_cash_flow="",
        )

        result.append(excel_row.to_dict())
        return result

    def generate_last_data(
        self,
        reimbursement_data: ReimbursementData,
        index: int,
        date: str,
        person_index: int,
    ) -> List[Dict[str, Any]]:
        """生成最后一行数据（贷方记录）"""
        result = []
        real_date = get_real_date(date)

        if not real_date:
            return result

        # 创建Excel行数据
        excel_row = ReimbursementExcelRowData(
            f_date=real_date.strftime("%Y-%m-%d"),
            f_year=real_date.strftime("%Y"),
            f_period=real_date.strftime("%m"),
            f_group_id="记",
            f_number=person_index,
            f_account_num=reimbursement_data.fee_code,
            f_account_name=reimbursement_data.fee_type,
            f_currency_num="RMB",
            f_currency_name="人民币",
            f_amount_for=reimbursement_data.amount,
            f_debit=0,
            f_credit=reimbursement_data.amount,
            f_preparer_id=DEFAULT_USERS["PREPARER"],
            f_checker_id="NONE",
            f_approve_id="NONE",
            f_cashier_id="NONE",
            f_handler="",
            f_settle_type_id="*",
            f_settle_no="",
            f_explanation=str(date) + str(reimbursement_data.summary or ""),
            f_quantity=0,
            f_measure_unit_id="*",
            f_unit_price=0,
            f_reference="",
            f_trans_date=real_date.strftime("%Y-%m-%d"),
            f_trans_no="",
            f_attachments=0,
            f_serial_num=1,
            f_object_name="",
            f_parameter="",
            f_exchange_rate=1,
            f_entry_id=index,
            f_item=reimbursement_data.department_project,
            f_posted=0,
            f_internal_ind="",
            f_cash_flow="",
        )

        result.append(excel_row.to_dict())
        return result

    def generate_last_row(
        self,
        top_info: ReimbursementTopInfo,
        base_data_item: ReimbursementData,
        index: int,
        total_amount: float,
        person_index: int,
        date: str,
    ) -> List[Dict[str, Any]]:
        """生成最后一行记录"""
        # 创建最后一行的报销数据，复制base_data_item的内容
        last_row_info = ReimbursementData(
            person=base_data_item.person,
            department=base_data_item.department,
            project=base_data_item.project,
            remark=base_data_item.remark,
            fee_type=top_info.bank,  # 使用top_info的bank作为fee_type
            fee_code=top_info.bank_code,  # 使用top_info的bank_code作为fee_code
            amount=total_amount,
            summary=top_info.summary,  # 使用top_info的summary
            department_project="",  # 部门+项目设为空
        )

        return self.generate_last_data(last_row_info, index, date, person_index)

    def process_reimbursement_data(
        self, top_infos: List[ReimbursementTopInfo], base_data: List[ReimbursementData]
    ) -> List[Dict[str, Any]]:
        """处理报销数据，生成最终的Excel数据"""
        result = []

        if not top_infos:
            return result

        date = top_infos[0].date
        person_index = 1

        for top_info in top_infos:
            index = 0
            total_amount = 0.0
            person = top_info.person

            # 处理每个人的报销数据
            for reimbursement in base_data:
                # 跳过不属于当前人员或空数据的记录
                if (
                    reimbursement.person is None
                    or reimbursement.person == ""
                    or person != reimbursement.person
                ):
                    continue

                # 处理金额
                amount = reimbursement.amount or 0
                reimbursement.amount = round_amount(amount, 2)
                total_amount += reimbursement.amount
                total_amount = round_amount(total_amount, 2)

                # 生成单条数据
                single_data = self.generate_excel_single_data(
                    reimbursement, index, date, person_index
                )
                result.extend(single_data)
                index += 1

            # 生成最后一行记录 - 与原始代码逻辑保持一致
            if person_index <= len(base_data):
                last_row = self.generate_last_row(
                    top_info,
                    base_data[person_index - 1],
                    index,
                    total_amount,
                    person_index,
                    date,
                )
                result.extend(last_row)

            person_index += 1

        return result
