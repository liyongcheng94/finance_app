"""
文件类型检测器 - 自动识别Excel文件是排单还是报销类型
"""

import openpyxl
import sys
import os
from typing import Optional


class FileTypeDetector:
    """文件类型检测器"""

    def __init__(self):
        self.workbook = None

    def detect_file_type(self, file_path: str) -> str:
        """
        检测文件类型
        参数: file_path - Excel文件路径
        返回: "payment" 或 "reimbursement"
        """
        try:
            # 首先通过文件名快速判断
            filename = os.path.basename(file_path).lower()
            if "报销" in filename or "reimbursement" in filename:
                print("通过文件名检测到报销文件")
                return "reimbursement"
            elif "排单" in filename or "payment" in filename:
                print("通过文件名检测到排单文件")
                return "payment"

            # 如果文件名无法判断，通过内容检测
            self.workbook = openpyxl.load_workbook(file_path, data_only=True)

            # 获取所有工作表名称
            sheet_names = self.workbook.sheetnames
            print(f"检测到工作表: {sheet_names}")

            # 检查工作表名称
            sheet_names_str = " ".join(sheet_names).lower()
            if "报销" in sheet_names_str or "reimbursement" in sheet_names_str:
                print("通过工作表名称检测到报销文件")
                return "reimbursement"

            # 检查内容
            return self._detect_by_content()

        except Exception as e:
            print(f"文件类型检测失败: {e}")
            # 默认返回排单类型
            return "payment"
        finally:
            if self.workbook:
                self.workbook.close()

    def _detect_by_content(self) -> str:
        """通过内容检测文件类型"""
        try:
            # 遍历所有工作表检查内容
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]

                # 检查前几行内容
                for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
                    if row:
                        row_text = " ".join([str(cell) for cell in row if cell])

                        # 报销文件特征关键词
                        reimbursement_keywords = [
                            "报销人",
                            "银行",
                            "FAmountFor",
                            "报销",
                            "部门代码",
                            "费用类型",
                        ]
                        if any(
                            keyword in row_text for keyword in reimbursement_keywords
                        ):
                            print(
                                f"通过内容检测到报销文件 (关键词: {row_text[:50]}...)"
                            )
                            return "reimbursement"

                        # 排单文件特征关键词
                        payment_keywords = [
                            "支付公司",
                            "供应商",
                            "排单",
                            "付款方式",
                            "户型",
                        ]
                        if any(keyword in row_text for keyword in payment_keywords):
                            print(
                                f"通过内容检测到排单文件 (关键词: {row_text[:50]}...)"
                            )
                            return "payment"

            # 如果都检测不到，默认为排单
            print("未检测到明确特征，默认为排单文件")
            return "payment"

        except Exception as e:
            print(f"内容检测失败: {e}")
            return "payment"  # 默认为排单

    def get_file_info(self, file_path: str) -> dict:
        """获取文件信息，用于调试"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            info = {
                "filename": os.path.basename(file_path),
                "sheet_names": workbook.sheetnames,
                "sheet_contents": {},
            }

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                content = []
                for i, row in enumerate(
                    sheet.iter_rows(min_row=1, max_row=5, values_only=True)
                ):
                    if i >= 5:  # 只获取前5行
                        break
                    content.append(
                        [str(cell) if cell else "" for cell in row[:10]]
                    )  # 只取前10列
                info["sheet_contents"][sheet_name] = content

            workbook.close()
            return info

        except Exception as e:
            return {"error": str(e)}


if __name__ == "__main__":
    # 测试代码
    detector = FileTypeDetector()

    # 测试文件路径
    test_file = input("请输入要测试的Excel文件路径: ")
    if test_file:
        file_type = detector.detect_file_type(test_file)
        print(f"检测结果: {file_type}")

        # 显示文件信息
        info = detector.get_file_info(test_file)
        print("文件信息:")
        print(f"文件名: {info.get('filename')}")
        print(f"工作表: {info.get('sheet_names')}")
